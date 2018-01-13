# -*- coding: utf-8 -*-

# Python Ver : Python 2.7.14
# Author : hwijinkim , https://github.com/jeonghoonkang
# last modified by 20171121

# Excel에서 원하는 list를 dic으로 바꾸기 위한 class

import json
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time
from __builtin__ import int

import sys
sys.path.insert(0, 'C:\Users\Be Pious\git\kdatahub\EE-data-analysis\lib')

import in_list

# only for sheet of 최종(LED+인버터)
column_ref = {
#     'company': ['E', 'company'],
#     'device': ['F', 'device_type'],
#     'building': ['M', 'building'],
#     'load': ['W', 'load'],
#     'led_serial': ['S', '_mds_id'],
#     'led_modem_serial': ['T', 'modem_num'],
#     'led_size': ['R', 'led_size'],
#     'inv_serial': ['AB', '_mds_id'],
#     'inv_modem_serial': ['AC', 'modem_num'],
#     'inv_size': ['Z', 'inv_size'],
    'sa_num': ['A', 'sa_num'],
    'business': ['B', 'business'],
    'yongdo': ['C', 'yongdo'],
    'seboo': ['D', 'seboo'],
    }

tag_ref = {
    # 건물용도(building)
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'bigmart', u'기타': 'etc',
    u'병원': 'hospital', u'전자대리점': 'electricmart', u'물류센터': 'boxmart',
    
    # 사업장구분(company)
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'biyoung', u'해당없음': 'nono',
    
    # 업종별-세부용도(business, detail)
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'fceramic',
    u'상용': 'sangyong', u'자동차부품': 'carpart', u'기계설비': 'machineEquipment', u'반도체': 'semiconductor', u'전자전기': 'electronicShock', u'의료': 'medical',
    
    # 품목(device_type)
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',
    
    # 부하(load)
    u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    u'터보냉동기': 'turbo_cold', u'냉각수펌프': 'pump_cold', u'냉온수기': 'cold_hot_dispensers', u'냉온수펌프':'cold_hot_pump', u'이에이치피':'ehp',
    
    #예외(값 안쓴것)
    np.NaN: 'none'
    }

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def thin_device(row):
    thin = {get_tag_ref('yongdo'): 'none'}
    if row[col2num(get_col_ref('sa_num'))].value != None:
#             thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
        thin[get_tag_ref('sa_num')] = str(row[col2num(get_col_ref('sa_num'))].value)
        thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
        thin[get_tag_ref('yongdo')] = conv2tag(row[col2num(get_col_ref('yongdo'))].value)
        thin[get_tag_ref('seboo')] = conv2tag(row[col2num(get_col_ref('seboo'))].value)
#             thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
#             thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
#             thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
#             thin[get_tag_ref('led_size')] = str(row[col2num(get_col_ref('led_size'))].value)
#             thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
#     else:
#         thin[get_tag_ref('sa_num')] = str(row[col2num(get_col_ref('business'))].value)
#         thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
#         thin[get_tag_ref('yongdo')] = conv2tag(row[col2num(get_col_ref('yongdo'))].value)
#         thin[get_tag_ref('seboo')] = conv2tag(row[col2num(get_col_ref('seboo'))].value)
#             thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
#             thin[get_tag_ref('sa_num')] = str(row[col2num(get_col_ref('sa_num'))].value)
#             thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
#             thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
#             thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
#             thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
#             thin[get_tag_ref('led_size')] = str(row[col2num(get_col_ref('led_size'))].value)
#             thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
#     elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
#         if row[col2num(get_col_ref('inv_size'))].value == None:
#             thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
#             thin[get_tag_ref('sa_num')] = str(row[col2num(get_col_ref('sa_num'))].value)
#             thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
#             thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
#             thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
#             thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
#             thin[get_tag_ref('inv_size')] = str(row[col2num(get_col_ref('inv_size'))].value)
#             thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
#         else:
#             thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
#             thin[get_tag_ref('sa_num')] = row[col2num(get_col_ref('sa_num'))].value
#             thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
#             thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
#             thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
#             thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
#             thin[get_tag_ref('inv_size')] = str(row[col2num(get_col_ref('inv_size'))].value)
#             thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':    
    led_list = []
    inv_list = []
    pure_led_list = []
    pure_inv_list = []
    led_gyo = []
    inv_gyo = []
    total_num = 1
    name = ''
    
    led_dict = "{"
    inv_dict = "{"

    wb = load_workbook(filename='160719_nangnanbang.xlsx', read_only=True)
    ws = wb[u'Sheet3']

    #1. extract tags from excel sheet by device type and size
#     for row in ws.iter_rows(min_row=2):
#         thin = thin_device(row)
#         led_list.append(thin)
#         elif thin[get_tag_ref('device')] == 'inverter':
#                 inv_list.append(thin)

#     for j in in_list.sanum_led_valid_list:
#         if in_list.all_sanum_dict[0].has_key(j):
#             for i in in_list.all_sanum_dict[0]:
#                 if i in in_list.ex_led_sanum_list:
#                     led_valid_list.append(i)
#             print '\n# led_valid_list count = %s' % len(led_valid_list)
#             print 'led_valid_list = %s\n' % led_valid_list
#     
#     print '<총 {0!r}개의 LED와 {1!r}개의 인버터 데이터 추출>\n'.format(len(led_list), len(inv_list))
    
#     for led in led_list:
#         if led['business'] == 'none':continue
#         pure_led_list.append(led)
#         led_dict += "'" + led['sa_num'] + "' : '" + led['business'] + "' , "
#     led_dict += '}\n'
    
    for j in in_list.excel_led_all_info_list:
        if in_list.all_sanum_dict[0].has_key(j['sa_num']):
            pure_led_list.append(j['modem_num'])
            led_dict += "'" + j['modem_num'] + "' : '" + in_list.all_sanum_dict[0][j['sa_num']] + "' , "
    led_dict += '}\n'
    
    print '# total_led = %s' % len(pure_led_list)
    print 'led_list = %s' % pure_led_list
    print '# total_led : %s' % len(led_dict)
    print 'led_dict = %s' % led_dict
    
    for j in in_list.excel_inv_all_info_list:
        if in_list.all_sanum_dict[0].has_key(j['sa_num']):
            pure_inv_list.append(j['modem_num'])
            inv_dict += "'" + j['modem_num'] + "' : '" + in_list.all_sanum_dict[0][j['sa_num']] + "' , "
    inv_dict += '}\n'
    
#     for inv in inv_list:
#         if inv['business'] == 'none':continue
#         pure_inv_list.append(inv)
#         inv_dict += "'" + inv['sa_num'] + "' : " + inv['business'] + "' , "
#     inv_dict += '}\n'
    
    print '# total_inv : %s' % len(pure_inv_list)
    print 'inv_list = %s' % pure_inv_list
    print '# total_inv : %s' % len(inv_dict)
    print 'inv_dict = %s' % inv_dict
    
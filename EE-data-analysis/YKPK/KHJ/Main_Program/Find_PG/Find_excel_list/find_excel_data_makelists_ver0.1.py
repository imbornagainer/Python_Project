# -*- coding: utf-8 -*-

# Author : hwijinkim , https://github.com/jeonghoonkang

# Excel에서 원하는 tag를 찾기 위한 class
# last modified by 20171117

import json
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time

# file.py
#import sys
#sys.path.insert(0, 'C:\Users\Be Pious\git\kdatahub\EE-data-analysis\lib')

#import in_list

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['W', 'load'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'], 
    'led_size': ['R', 'size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'size'],
    'business': ['AL', 'business'],
    }

tag_ref = {
    # 건물용도(building)
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'bigmart', u'기타': 'etc',
    u'병원': 'hospital', u'전자대리점': 'electricmart',
    
    # 사업장구분(company)
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'biyoung', u'해당없음': 'nono',
    
    # 업종별-세부용도(business, detail)
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'fceramic',
    u'상용': 'sangyong',
    
    # 품목(device_type)
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',
    
    # 부하(load)
    u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    
    #예외(값 안쓴것)
    np.NaN: 'none'
    }

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def thin_device(row):

    thin = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        if row[col2num(get_col_ref('led_size'))].value == None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = 'none'
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
        else:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        if row[col2num(get_col_ref('inv_size'))].value == None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = 'none'
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
        else:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin
        
# Lists 기록함수
def Write_Lists(device_list, building_list, load_list, file_nm, device_tp, building_tp, load_tp):
    f = open(file_nm, 'w')    # 경로설정 / 덮어쓰기
    f.write('%s_cnt = %s\n' % (device_tp, len(device_list)))
    f.write('%s_list = %s\n' % (device_tp, device_list))
    f.write('%s_cnt = %s\n' % (building_tp, len(building_list)))
    f.write('%s_list = %s\n' % (building_tp, building_list))
    f.write('%s_cnt = %s\n' % (load_tp, len(load_list)))
    f.write('%s_list = %s\n' % (load_tp, load_list))
    f.write('\n')
    f.close()

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':    
    device_list = []
    building_list = []
    load_list = []
    pure_device_list = []
    pure_building_list = []
    pure_load_list = []
    device_gyo = []
    building_gyo = []
    load_gyo = []    
    name = ''
    
    # 1. 설정값 입력 - 시작
    wb = load_workbook(filename='(EE)list_002.xlsx', read_only=True)    # Excel file명 입력 (첫번째 인자)
    ws = wb[u'최종(LED+인버터)']                                        # 사용할 엑셀 Sheets 이름 입력
    
    device_tp   = 'inverter'                        # 장비종류 설정
    building_tp = 'factory'                         # 빌딩종류 설정
    load_tp     = 'blower'                          # 로드종류 설정
    file_nm     = '171117_tag_lists_in_excel.py'    # 리스트 입력할 이름명 설정(확장자까지 입력)
    # 설정값 입력   - 끝
    
    #1. extract tags from excel sheet by device type and size
    for row in ws.iter_rows(min_row=4):
        thin = thin_device(row)        
        if 'device_type' in thin.keys():
            if thin[get_tag_ref('device')] == device_tp:
                device_list.append(thin)
        if 'building' in thin.keys():
            if thin[get_tag_ref('building')] == building_tp:
                building_list.append(thin)
        if 'load' in thin.keys():
            if thin[get_tag_ref('load')] == load_tp:
                load_list.append(thin)
    print '<총 {0!r}개의 {1!r}, {2!r}개의 {3!r}, {4!r}개의 {5!r} 데이터 추출>\n'.format(len(device_list), device_tp, len(building_list), building_tp, len(load_list), load_tp)
    
    for i in range(len(device_list)):
        pure_device_list.append(device_list[i]['modem_num'])
    print 'total_%s : %s' % (device_tp, len(pure_device_list))
    print '%s_list = %s\n' % (device_tp, pure_device_list)
    
    for j in range(len(building_list)):
        pure_building_list.append(building_list[j]['modem_num'])
    print 'total_%s : %s' % (building_tp, len(pure_building_list))
    print '%s_list = %s\n' % (building_tp, pure_building_list)
     
    for k in range(len(load_list)):
        pure_load_list.append(load_list[k]['modem_num'])
    print 'total_%s : %s' % (load_tp, len(pure_load_list))
    print '%s_list = %s\n' % (load_tp, pure_load_list)

    Write_Lists(pure_device_list, pure_building_list, pure_load_list, file_nm, device_tp, building_tp, load_tp)    
    
#     교집합인 것들만 lists 출력    - 추가할지는 추후에 결정
#     for i in led_excel:
#         if i in pure_led_list:
#             led_gyo.append(i)
#     print 'led_gyo_cnt = %s' % len(led_gyo)
#     print 'led_gyo = %s' % led_gyo
#     
#     for j in inv_excel:
#         if j in pure_inv_list:
#             inv_gyo.append(i)
#     print 'inv_gyo_cnt = %s' % len(inv_gyo)
#     print 'inv_gyo = %s\n' % inv_gyo
#     
#     for j in inv_excel:
#         if j in pure_inv_list:
#             inv_gyo.append(i)
#     print 'inv_gyo_cnt = %s' % len(inv_gyo)
#     print 'inv_gyo = %s\n' % inv_gyo

# -*- coding: utf-8 -*-

import json
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time
from __builtin__ import int

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['W', 'load'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'],
    'led_size': ['R', 'led_size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'inv_size'],
    'business': ['AL', 'business'],
    }

tag_ref = {
    # 건물용도(building)
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'bigmart', u'기타': 'etc',
    u'병원': 'hospital', u'전자대리점': 'electricmart',
    
    # 사업장구분(company)
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'biyoung', u'해당없음': 'nono',
    
    # 업종별-세부용도(business, detail)
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'fceramic',
    u'상용': 'sangyong',
    
    # 품목(device_type)
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',
    
    # 부하(load)
    u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    u'터보냉동기': 'turbo_cold', u'냉각수펌프': 'pump_cold', u'냉온수기': 'cold_hot_dispensers', u'냉온수펌프':'cold_hot_pump', u'이에이치피':'ehp',
    
    #예외(값 안쓴것)
    np.NaN: 'none'
    }

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def thin_device(row):
    thin = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        if row[col2num(get_col_ref('led_size'))].value == None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = 'none'
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('led_size')] = str(row[col2num(get_col_ref('led_size'))].value)
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
        else:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)            
            thin[get_tag_ref('led_size')] = str(row[col2num(get_col_ref('led_size'))].value)
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        if row[col2num(get_col_ref('inv_size'))].value == None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = 'none'
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('inv_size')] = str(row[col2num(get_col_ref('inv_size'))].value)
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
        else:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('inv_size')] = str(row[col2num(get_col_ref('inv_size'))].value)
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin
        
# Lists 기록함수
def Write_Lists(led_list, inv_list):
    f = open('excel_lists_171108_1.py', 'w')    # 경로설정 / 덮어쓰기
    f.write('led_list = %s\n' % led_list)
    f.write('inv_list = %s\n' % inv_list)
    f.write('\n')
    f.close()

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':    
    led_list = []
    inv_list = []
    pure_led_list = []
    pure_inv_list = []
    led_gyo = []
    inv_gyo = []
    total_num = 1
    name = ''
    
    led_dict = "{"
    inv_dict = "{"

    wb = load_workbook(filename='(EE)list_002.xlsx', read_only=True)
    ws = wb[u'최종(LED+인버터)']
    
    #led cnt = 314
    led_excel = ['01222799986', '01220800574', '01224230529', '01222714543', '01224230467', '01223239869', '01224230576', '01222679989', '01222746396', '01222746350', '01224230549', '01224224135', '01224273071', '01224272966', '01224230577', '01224230493', '01224230501', '01224230526', '01221386983', '01223239875', '01224230542', '01224230434', '01224230480', '01224230484', '01224230485', '01224230470', '01223192247', '01222745462', '01221574357', '01222742200', '01222754504', '01222739929', '01224230571', '01222674503', '01224230527', '01224230569', '01222669986', '01223187397', '01223209886', '01222764501', '01222742912', '01224272962', '01222745935', '01222745933', '01222745947', '01222742969', '01222746463', '01223239874', '01223239873', '01223187431', '01222674501', '01220684147', '01222694501', '01222734501', '01223187419', '01223188270', '01223189537', '01222769929', '01221574356', '01224230509', '01224230472', '01224230478', '01222746390', '01222745510', '01222719929', '01221574359', '01222746366', '01222779929', '01223239865', '01223219128', '01224230518', '01223187436', '01223188269', '01222754503', '01222742885', '01222745938', '01222724503', '01222704504', '01222742752', '01222694503', '01224230546', '01221341991', '01224224149', '01220800623', '01220435025', '01224230511', '01224230533', '01224230451', '01223209883', '01223209884', '01223219126', '01223209881', '01223192283', '01222654504', '01222755541', '01222766541', '01222709541', '01222744754', '01222719541', '01222739541', '01222587543', '01222666541', '01223192386', '01222714503', '01224230446', '01224230532', '01222667543', '01222764503', '01222586543', '01222742956', '01222684501', '01222699984', '01224230558', '01224230495', '01222739987', '01221961213', '01223187386', '01223192395', '01224273086', '01222567452', '01222587452', '01221566970', '01222714504', '01222664504', '01222676543', '01222647452', '01222744504', '01222749989', '01222746383', '01222776543', '01222745635', '01222597452', '01224230436', '01222758870', '01222779987', '01224273013', '01222745527', '01222745458', '01222664501', '01222779986', '01224230496', '01222767541', '01222742927', '01222689984', '01222687452', '01222742760', '01224230525', '01222797541', '01224230439', '01224272937', '01224230449', '01224230457', '01224230506', '01222746381', '01224230477', '01224273037', '01223192392', '01224230491', '01221421994', '01221441994', '01222749929', '01224230551', '01224273073', '01224230473', '01224230530', '01223239872', '01224230488', '01224273039', '01224273091', '01224230570', '01224272955', '01224230483', '01224230513', '01224230482', '01223239867', '01224273038', '01224230490', '01222704501', '01223239866', '01222754501', '01222709929', '01222748870', '01224230528', '01222683542', '01222664503', '01222654503', '01222674504', '01222746482', '01222719989', '01221448178', '01222744503', '01223192284', '01223187417', '01223187414', '01222759929', '01221243882', '01221361991', '01223192328', '01221371991', '01221474413', '01221273883', '01222654501', '01222799985', '01221253884', '01220800593', '01222742949', '01222746459', '01222684503', '01224230460', '01224230565', '01224230465', '01221388068', '01221391442', '01221391244', '01221391199', '01221391155', '01224230476', '01221574346', '01221567020', '01222765541', '01222734504', '01221566961', '01221566960', '01223189591', '01221566971', '01223187435', '01221574354', '01223188887', '01221574324', '01221567011', '01222742946', '01222636543', '01222659984', '01222563542', '01224230564', '01224230568', '01220800748', '01222744780', '01224230541', '01224230539', '01224230534', '01222745479', '01222729989', '01222577452', '01222746397', '01222779985', '01222769989', '01222729987', '01224230487', '01223209889', '01223209891', '01223209888', '01224230489', '01224230504', '01224230486', '01224230469', '01224230510', '01222746461', '01224230443', '01222796541', '01224230512', '01222742798', '01222746421', '01222742888', '01222788870', '01222746472', '01222597541', '01222769542', '01222616541', '01224230437', '01223209890', '01222724501', '01221283886', '01222564542', '01222742908', '01222745934', '01222745495', '01222745859', '01224230516', '01224230500', '01223239868', '01224230575', '01224230573', '01222769984', '01223187402', '01223187398', '01223187384', '01223187403', '01224272944', '01221429366', '01221429367', '01221429369', '01221471994', '01221474403', '01221474402', '01221474406', '01221474407', '01221341993', '01221371993', '01223192410', '01223192453', '01221243887', '01223192393', '01223192394', '01223187395', '01224230458', '01224230459', '01222746380', '01222794503', '01220800701', '01222747452', '01221429368', '01221390254', '01222789986']
    #inv cnt = 50
    inv_excel = ['01224230502', '01224273014', '01221392626', '01221392897', '01224272943', '01224273042', '01224230567', '01224273017', '01223187286', '01223180993', '01220553054', '01224273061', '01224273022', '01221392027', '01221389654', '01221390900', '01221387086', '01221387826', '01221399481', '01221393131', '01221393546', '01224273046', '01224273052', '01223239879', '01223239880', '01223239878', '01223239876', '01223192368', '01223192373', '01224230548', '01223181052', '01224273085', '01224273018', '01224273096', '01224273087', '01223180966', '01223180976', '01224230447', '01224230455', '01224230466', '01224230441', '01224230521', '01224273045', '01224272958', '01224230559', '01224273026', '01224273079', '01224272971', '01220800675', '01224272988']

    #1. extract tags from excel sheet by device type and size
    for row in ws.iter_rows(min_row=4):
        thin = thin_device(row)
        if thin[get_tag_ref('device')] == 'led':
            led_list.append(thin)
        elif thin[get_tag_ref('device')] == 'inverter':
            inv_list.append(thin)
    print '<총 {0!r}개의 LED와 {1!r}개의 인버터 데이터 추출>\n'.format(len(led_list), len(inv_list))
    
    for led in led_list:
        led_dict += "'" + led['modem_num'] + "':" + led['led_size'] + " , "
    led_dict += '}'
    print 'total_led : %s' % len(led_dict)
    print 'led_dict = %s' % led_dict
    
    for inv in inv_list:
        inv_dict += "'" + inv['modem_num'] + "':" + inv['led_size'] + " , "
    print 'total_inv : %s' % len(inv_dict)
    print 'inv_dict = %s' % inv_dict
    
#     for i in led_excel:
#         if i in pure_led_list:
#             led_gyo.append(i)
#     print 'led_gyo_cnt = %s' % len(led_gyo)
#     print 'led_gyo = %s' % led_gyo
#     
#     for j in inv_excel:
#         if j in pure_inv_list:
#             inv_gyo.append(i)
#     print 'inv_gyo_cnt = %s' % len(inv_gyo)
#     print 'inv_gyo = %s' % inv_gyo

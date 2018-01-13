# -*- coding: utf-8 -*-

import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import json
import socket
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time
import gzip
import shutil
import sys
sys.path.insert(0, 'C:\Users\Be Pious\git\kdatahub\EE-data-analysis\lib')
sys.path.insert(0, '/home/bornagain/kdatahub/EE-data-analysis/lib/tools')
sys.path.insert(0, '/home/bornagain/kdatahub/EE-data-analysis/lib')

import in_list

USE_HTTP_API_PUT = False

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['H', 'load'],
    'modem_num': ['A', 'modem_num'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'],
    'led_size': ['R', 'size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'size'],
    'detail': ['AL', 'detail'],
    }

tag_ref = {
    # 건물용도(building)
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'bigmart', u'기타': 'etc',
    u'병원': 'hospital', u'전자대리점': 'electricmart',

    # 사업장구분(company)
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'biyoung', u'해당없음': 'nono',

    # 업종별-세부용도(business, detail)
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'fceramic',
    u'상용': 'sangyong', u'물류센터': 'box_center',

    # 품목(device_type)
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',

    # 부하(load)
    u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    u'터보냉동기': 'turbo_cold', u'냉각수펌프': 'pump_cold', u'냉온수기': 'cold_hot_dispensers', u'냉온수펌프':'cold_hot_pump', u'이에이치피':'ehp',

    #예외(값 안쓴것)
    np.NaN: 'none'
    }

def thin_device(row):
    thin = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        #if row[col2num(get_col_ref('led_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('detail')] = conv2tag(row[col2num(get_col_ref('detail'))].value)
            #thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        #if row[col2num(get_col_ref('inv_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            thin[get_tag_ref('detail')] = conv2tag(row[col2num(get_col_ref('detail'))].value)
            #thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):    
    if val_str in tag_ref:        
        return tag_ref[val_str]        
    else:
        return 'ehp'        

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def pack_dps(metric, dps, tags):
    pack = []
    tag = {'metric': metric, 'tags': tags}
    
    for dp in dps:
        sdp = copy.copy(tag)
        sdp['timestamp'] = int(dp.encode('utf8'))
        sdp['value'] = dps[dp]
        pack.append(sdp)
    return pack

def requests_retry_session(
    retries=3,
    backoff_factor=0.3,
    status_forcelist=(500, 502, 504),
    session=None,
):
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

def tsdb_query(query, start, end, metric, modem_num, holiday):
    param = {}
    if start: param['start'] = start
    if end: param['end'] = end
    param['m'] = metric + '{' + 'holiday={1!s},modem_num={0!s}'.format(modem_num, holiday) + '}'

    if 1:
        response = requests.get(query, params=param)
        if response.ok:
            return response.json()
        else:
            print 'request fails'
            return []
    else:
        try:
            response = requests_retry_session().get(query, params=param, timeout=5)
        except Exception as x:
            print 'requests timeout'
        else:
            if response.ok:
                return response.json()
        return []

def tsdb_put(put, metric, dps, tags, file=None):
    packed_data = pack_dps(metric, dps, tags)
    for i in xrange(0, len(packed_data), 8):
        tmp = json.dumps(packed_data[i:i+8])
        if 0:
            response = requests.post(put, data=tmp)
        if response.text: print response.text
        else:
            try:
                response = requests_retry_session().post(put, data=tmp, timeout=5)
            except Exception as x:
                print 'requests timeout'
            else:
                if response.text: print response.text
    return

def tsdb_put_telnet(host, port, metric, dps, tags, file=None):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    retry = 9999
    while 1:
        try:
            sock.connect((host, port))
        except:
            print 'socket error: ' + host
            time.sleep(5)
            print 'time sleep 5second'
            if retry > 0:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                continue
            else:
                print 'skip this dps'
                return
        else:
            break
    send = ''
    count = 0
    packed_data = pack_dps(metric, dps, tags)
    for dp in packed_data:
        if (dp['value'] != 0 and dp['value'] != None and dp['value'] > 0.19 and dp['value'] <= 100.0):
            if len(send) >= 1024:
                sock.sendall(send)
                if file: file.write(send)
                send = ''
            send += 'put'
            send += ' {0!s}'.format(dp['metric'])
            send += ' {0!r}'.format(dp['timestamp'])
            send += ' {0!r}'.format(dp['value'])
            if 'tags' in dp:
                for key in dp['tags']:
                    send += ' {0!s}'.format(key) + '={0!s}'.format(dp['tags'] [key])
            send += '\n'            
            count += 1
        else:
            global chk_number
            chk_number = (chk_number + 1)
            Chk_Outlier(chk_number, dp['tags']['modem_num'], dp['tags']['_mds_id'], dp['timestamp'], dp['value'])
    if len(send) > 0:
        sock.sendall(send)
        if file: file.write(send)
        send = ''
    sock.close()
    return

# outlier log 기록함수
# open의 경로명을 설정해주어야 한다.
def Chk_Outlier(chk_num, adj_modem_num,mds_id,unix_time, value):
    if chk_num == 1:
        f = open('outlier_lists_180110_v2.txt', 'w')    # 경로설정 / 덮어쓰기
        f.write('%d:\n' % chk_num)
        f.write('modem_num : %s\n' % adj_modem_num)
        f.write('_mds_id   : %s\n' % mds_id)
        f.write('unix_time : %s\n' % ts2datetime(unix_time))
        f.write('value     : %s\n' % value)
        f.write('\n')
        f.close()
    else:
        f = open('outlier_lists_180110_v2.txt', 'a')    # 경로설정 / 이어쓰기
        f.write('%d:\n' % chk_num)
        f.write('modem_num : %s\n' % adj_modem_num)
        f.write('_mds_id   : %s\n' % mds_id)
        f.write('unix_time : %s\n' % ts2datetime(unix_time))
        f.write('value     : %s\n' % value)
        f.write('\n')
        f.close()

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':
    host  = '125.140.110.217'
    port  = 4242
    query = 'http://{0!s}:{1!r}/api/query'.format(host, port)
    put   = 'http://{0!s}:{1!r}/api/put?summary'.format(host, port)
    first = '2016/07/01-00:00:00'
    last  = '2017/07/01-00:00:00'
    metric_in  = 'none:rc04_simple_data_v3'
    chk_number = 0
    metric_out = 'rc05_excel_copy_tag_v8'
    out_file = open(metric_out, 'w')
    led_list = []
    inv_list = []
    led_list_all = []
    inv_list_all = []

    wb = load_workbook(filename='180110_EE_list.xlsx', read_only=True)
    ws = wb[u'최종(LED+인버터)']

    #1. extract tags from excel sheet by device type and size
    for row in ws.iter_rows(min_row=4):
        thin = thin_device(row)
        if thin[get_tag_ref('device')] == 'led':
            led_list.append(thin)
        elif thin[get_tag_ref('device')] == 'inverter':
            inv_list.append(thin)

    print '<총 {0!r}개의 LED와 {1!r}개의 인버터 데이터 추출>\n'.format(len(led_list), len(inv_list))

    led_count = 0
    inv_count = 0
    start_time = datetime.datetime.now()

    #2. query data points from tsdb using modem number
    #3. add tags and put data points to tsdb
    for led in led_list:
        #print led
        #iterate every days and must be started w/ start of the day
        print led['modem_num']
        led_list_all.append(led['modem_num'])
#         tmp_time = datetime.datetime.now()
#         end = first
#         while is_past(end, last):
#             start = end
#             end = add_time(start, days=1)
#             led['holiday'] = is_weekend(start)
#             query_result = tsdb_query(query, start, end, metric_in, led[get_tag_ref('led_modem_serial')], led['holiday'])
#             for group in query_result:
#                 if USE_HTTP_API_PUT:
#                     tsdb_put(put, metric_out, group['dps'], led, out_file)
#                 else:
#                     tsdb_put_telnet(host, port, metric_out, group['dps'], led, out_file)
#         led_count += 1
#         print 'elapsed time: {0!s}, {1!s}'.format(datetime.datetime.now() - tmp_time, datetime.datetime.now() - start_time)
#         print '<총 {0!r}/{1!r}개의 LED와 {2!r}/{3!r}개의 인버터 데이터 처리>\n'.format(led_count, len(led_list), inv_count, len(inv_list))

    for inv in inv_list:
        #print inv
        inv_list_all.append(inv['modem_num'])
#         tmp_time = datetime.datetime.now()
#         end = first
#         while is_past(end, last):
#             start = end
#             end = add_time(start, days=1)
#             inv['holiday'] = is_weekend(start)
#             query_result = tsdb_query(query, start, end, metric_in, inv[get_tag_ref('inv_modem_serial')], inv['holiday'])
#             for group in query_result:
#                 if USE_HTTP_API_PUT:
#                     tsdb_put(put, metric_out, group['dps'], inv, out_file)
#                 else:
#                     tsdb_put_telnet(host, port, metric_out, group['dps'], inv, out_file)
#         inv_count += 1
#         print 'elapsed time: {0!s}, {1!s}'.format(datetime.datetime.now() - tmp_time, datetime.datetime.now() - start_time)
#         print '<총 {0!r}/{1!r}개의 LED와 {2!r}/{3!r}개의 인버터 데이터 처리>\n'.format(led_count, len(led_list), inv_count, len(inv_list))
# 
#     print 'total elapsed time: {0!s}'.format(datetime.datetime.now() - start_time)

    print led_list_all
    print inv_list_all

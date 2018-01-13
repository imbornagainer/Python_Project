# -*- coding: utf-8 -*-
# Author : hwijinkim , https://github.com/jeonghoonkang

# add_Tag를 위한 Class 
# last modified by 20171018
# tsdb_query NONE return 예외처리 추가

import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import json
import socket
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time

USE_HTTP_API_PUT = False

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['W', 'load'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'],
    'led_size': ['R', 'size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'size'],
    'business': ['AZ', 'business'],
    }

tag_ref = {
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'mart', u'기타': 'etc',  # 건물용도
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'etc', u'해당없음': 'etc',
    u'병원': 'hospital',  # 사업장구분
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'ceramic',
    u'상용': 'sangyong',  # 업종별-세부용도
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',  # 품목
    np.NaN: 'none', u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    }

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def thin_device(row):
    thin = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        if row[col2num(get_col_ref('led_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            #thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        if row[col2num(get_col_ref('inv_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            #thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin

def pack_dps(metric, dps, tags):
    pack = []
    tag = {'metric': metric, 'tags': tags}
    for dp in dps:
        sdp = copy.copy(tag)
        sdp['timestamp'] = int(dp.encode('utf8'))
        sdp['value'] = dps[dp]
        #print sdp
        pack.append(sdp)
    #print pack
    return pack

def requests_retry_session(
    retries=3,
    backoff_factor=0.3,
    status_forcelist=(500, 502, 504),
    session=None,
):
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    print adapter
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    print session.mount('https://', adapter)
    return session

def tsdb_query(query, start, end, metric, modem_num, holiday):
    param = {}
    if start: param['start'] = start
    if end: param['end'] = end
    param['m'] = metric + '{' + 'holiday={1!s},modem_num={0!s}'.format(modem_num, holiday) + '}'
    #print param
    if 0:
        response = requests.get(query, params=param)
        if response.ok:
            #print response.text
            return response.json()
        else:
            print 'request fails'
            return []
    else:
        try:
            response = requests_retry_session().get(query, params=param, timeout=5)
        except Exception as x:
            print 'requests timeout'
        else:
            if response.ok:
                # print response.text
                return response.json()
#         finally:
#             print 'request fails'
#             return []

def tsdb_put(put, metric, dps, tags):
    packed_data = pack_dps(metric, dps, tags)
    for i in xrange(0, len(packed_data), 8):
        tmp = json.dumps(packed_data[i:i+8])
        print tmp
        response = requests.post(put, data=tmp)
        if response.text: print response.text
    return

def tsdb_put_telnet(metric, dps, tags):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.connect((host, port))
    send = ''
    count = 0
    print 'put dps: {0} requested...'.format(len(dps))
    packed_data = pack_dps(metric, dps, tags)
    print 'put dps: {0} packed...'.format(len(packed_data))
    for dp in packed_data:
        if len(send) >= 1024:
            #print send + '\n\n'
            sock.sendall(send)
            send = ''
        send += 'put'
        send += ' {0!s}'.format(dp['metric'])
        send += ' {0!r}'.format(dp['timestamp'])
        send += ' {0!r}'.format(dp['value'])
        if 'tags' in dp:
            for key in dp['tags']:
                send += ' {0!s}'.format(key) + '={0!s}'.format(dp['tags'] [key])
        send += '\n'
        count += 1
    if len(send) > 0:
        #print send + '\n\n'
        sock.sendall(send)
        send = ''
    sock.close()
    print 'put dps: {0} processed...'.format(count)
    return

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    print datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':
    host  = '125.140.110.217'
    port  = 4242
    query = 'http://{0!s}:{1!r}/api/query'.format(host, port)
    put   = 'http://{0!s}:{1!r}/api/put?summary'.format(host, port)
    first = '2016/07/01-00:00:00'
    last  = '2017/07/01-00:00:00'
    #orgin_metric_in = none:rc04_simple_data_v3
    metric_in  = 'none:rc04_simple_data_v3'
    #metric_in  = 'sum:rc04_simple_data_v3'
    #orgin_metric_out = UltimateInfestation_v4
    metric_out = 'rc03_add_led_tag_v11'
    #metric_out = 'RazaTheChained_v2'
    led_list = []
    inv_list = []

    wb = load_workbook(filename='ref_list.xlsx', read_only=True)
    ws = wb[u'최종(LED+인버터)']

    #1. extract tags from excel sheet by device type and size
    for row in ws.iter_rows(min_row=4):
        thin = thin_device(row)
        if thin[get_tag_ref('device')] == 'led':
            led_list.append(thin)
        elif thin[get_tag_ref('device')] == 'inverter':
            inv_list.append(thin)

    print '<총 {0!r}개의 LED와 {1!r}개의 인버터 데이터 추출>\n'.format(len(led_list), len(inv_list))

    led_count = 0
    inv_count = 0

    #2. query data points from tsdb using modem number
    #3. add tags and put data points to tsdb
    for led in led_list:
        # print led
        #iterate every days and must be started w/ start of the day
        end = first
        while is_past(end, last):
            start = end
            end = add_time(start, days=1)
            led['holiday'] = is_weekend(start)
            
            query_result = tsdb_query(query, start, end, metric_in, led[get_tag_ref('led_modem_serial')], led['holiday'])
            if query_result == None:continue #if no data, go to the next iteration
            for group in query_result:
                if 1:
                    tsdb_put(put, metric_out, group['dps'], led)
                else:
                    tsdb_put_telnet(metric_out, group['dps'], led)
        led_count += 1
        print '<총 {0!r}/{1!r}개의 LED와 {2!r}/{3!r}개의 인버터 데이터 처리>\n'.format(led_count, len(led_list), inv_count, len(inv_list))

    for inv in inv_list:
        print inv
        end = first
        while is_past(end, last):
            start = end
            end = add_time(start, days=1)
            inv['holiday'] = is_weekend(start)
            query_result = tsdb_query(query, start, end, metric_in, inv[get_tag_ref('inv_modem_serial')], inv['holiday'])
            if query_result == None:continue #if no data, go to the next iteration
            for group in query_result:
                if 0:
                    tsdb_put(put, metric_out, group['dps'], inv)
                else:
                    tsdb_put_telnet(metric_out, group['dps'], inv)
        inv_count += 1
        print '<총 {0!r}/{1!r}개의 LED와 {2!r}/{3!r}개의 인버터 데이터 처리>\n'.format(led_count, len(led_list), inv_count, len(inv_list))
    exit()

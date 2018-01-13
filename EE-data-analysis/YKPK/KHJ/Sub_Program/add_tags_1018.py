# -*- coding: utf-8 -*-

import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import json
import socket
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time
import gzip
import shutil
import openpyxl

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['W', 'load'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'],
    'led_size': ['R', 'size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'size'],
    'business': ['AZ', 'business'],
    }

tag_ref = {
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'mart', u'기타': 'etc',  # 건물용도
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'etc', u'해당없음': 'etc',
    u'병원': 'hospital',  # 사업장구분
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'ceramic',
    u'상용': 'sangyong',  # 업종별-세부용도
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',  # 품목
    np.NaN: 'none', u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    }

# 엑셀의 column값(알파벳)을 숫자로 - ex) T -> 20
def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1 # Column - 1 값이 원하는 값(5 출력시 F(6) Column의 값이 출력)

# 엑셀의 column값(알파벳)을 value로 - tag key값이 있으면 value값 return - ex) 공장 -> factory
def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

# 입력받은 key값이 있으면 list key값 return
def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

# 입력받은 key값이 있으면 list value값 return
def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def find_device(row):
    f_info = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        if row[col2num(get_col_ref('led_size'))].value != None:
            f_info[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            f_info[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            f_info[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            f_info[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            f_info[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            f_info[get_tag_ref('business')] = 'none'
            f_info[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        if row[col2num(get_col_ref('inv_size'))].value != None:
            f_info[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            f_info[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            f_info[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            f_info[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            f_info[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            f_info[get_tag_ref('business')] = 'none'
            f_info[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return f_info

def tsdb_query(query, start, end, metric, modem_num, holiday):
    param = {}
    if start: param['start'] = start
    if end: param['end'] = end
    param['m'] = metric + '{' + 'holiday={1!s},modem_num={0!s}'.format(modem_num, holiday) + '}'
    #print param['m']
    #print "param : %s" % param
    if 1:
        response = requests.get(query, params=param)
        if response.ok:
            #print "response.text : %s" % response.text
            #print "response.json : %s" % response.json
            return response.json()
        else:
            print 'request fails'
            return []
    else:
        try:
            response = requests_retry_session().get(query, params=param, timeout=5)
        except Exception as x:
            print 'requests timeout'
        else:
            if response.ok:
                #print response.text
                return response.json()
        return "end!"

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    
def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

if __name__ == '__main__':
    host  = '125.140.110.217'
    port  = 4242
    query = 'http://{0!s}:{1!r}/api/query'.format(host, port)
    put   = 'http://{0!s}:{1!r}/api/put?summary'.format(host, port)
    first = '2016/07/01-00:00:00'
    last  = '2017/07/01-00:00:00'
    metric_in  = 'none:rc04_simple_data_v3'
    #metric_in  = 'sum:rc04_simple_data_v3'
    metric_out = 'rc03_add_led_tag_v12'
    #metric_out = 'RazaTheChained_v4'
    out_file = open(metric_out, 'w')
    led_list = []
    inv_list = []
    chk_num = 0
    
    # 해당 문서의 sheet 탐색 - get_sheet_names()
    # 해당 문서의 다수 sheet 접근 - get_sheet_by_name()
    l_w = load_workbook('ref_list_kjh.xlsx')    
    l_ws = l_w[u'최종(LED+인버터)']
    
    for row in l_ws.iter_rows(min_row=4, max_row=539):   # 행을 기준으로 iter(반복문)
        if 0:
            for cell in row:
                print "cell : %s" % (cell)
                print "cell.value : %s" % (cell.value)
        else:
            finded = find_device(row)
            if finded[get_tag_ref('device')] == 'led':
                led_list.append(finded)
            elif finded[get_tag_ref('device')] == 'inverter':
                inv_list.append(finded)
    
    print "Load Result \nled_list count : %s \ninv_list count : %s \nsum count : %s" % (len(led_list), len(inv_list), len(led_list)+len(inv_list))    
    
    for led in led_list:
        print "led_list 실행 횟수 : %d\n" % (chk_num+1)
        end = first     # while문을 위한 변수 (1day씩 last년까지 반복하기 위해서)
        while(end != last):
            start = end
            led['holiday'] = is_weekend(start)
            end = add_time(start, days=1)
            query_result = tsdb_query(query, start, end, metric_in, led[get_tag_ref('led_modem_serial')], led['holiday'])
            chk_num += 1
            #print "query_result : " % led
            print "query_result : %s" % query_result
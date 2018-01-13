# -*- coding: utf-8 -*-

import json
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time

# only for sheet of 최종(LED+인버터)
column_ref = {
    'company': ['E', 'company'],
    'device': ['F', 'device_type'],
    'building': ['M', 'building'],
    'load': ['W', 'load'],
    'led_serial': ['S', '_mds_id'],
    'led_modem_serial': ['T', 'modem_num'],
    'led_size': ['R', 'size'],
    'inv_serial': ['AB', '_mds_id'],
    'inv_modem_serial': ['AC', 'modem_num'],
    'inv_size': ['Z', 'size'],
    'business': ['AZ', 'business'],
    }

tag_ref = {
    u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'mart', u'기타': 'etc',  # 건물용도
    u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'etc', u'해당없음': 'etc',
    u'병원': 'hospital', u'전자대리점': 'mart', # 사업장구분
    u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
    u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'ceramic',
    u'상용': 'sangyong',  # 업종별-세부용도
    u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',  # 품목
    np.NaN: 'none', u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
    }

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num - 1

def conv2tag(val_str):
    if val_str in tag_ref:
        return tag_ref[val_str]
    else:
        return 'none'

def get_col_ref(key):
    if key in column_ref:
        return column_ref[key][0]
    else:
        return 'ZZZ'

def get_tag_ref(key):
    if key in column_ref:
        return column_ref[key][1]
    else:
        return 'unknown'

def thin_device(row):
    thin = {get_tag_ref('device'): 'none'}
    if row[col2num(get_col_ref('led_modem_serial'))].value != None:
        if row[col2num(get_col_ref('led_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('led_serial')] = row[col2num(get_col_ref('led_serial'))].value.encode('utf8')
            thin[get_tag_ref('led_modem_serial')] = row[col2num(get_col_ref('led_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            #thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    elif row[col2num(get_col_ref('inv_modem_serial'))].value != None:
        if row[col2num(get_col_ref('inv_size'))].value != None:
            thin[get_tag_ref('device')] = conv2tag(row[col2num(get_col_ref('device'))].value)
            thin[get_tag_ref('inv_serial')] = row[col2num(get_col_ref('inv_serial'))].value.encode('utf8')
            thin[get_tag_ref('inv_modem_serial')] = row[col2num(get_col_ref('inv_modem_serial'))].value.encode('utf8')
            thin[get_tag_ref('company')] = conv2tag(row[col2num(get_col_ref('company'))].value)
            thin[get_tag_ref('building')] = conv2tag(row[col2num(get_col_ref('building'))].value)
            #thin[get_tag_ref('business')] = conv2tag(row[col2num(get_col_ref('business'))].value)
            thin[get_tag_ref('business')] = 'none'
            thin[get_tag_ref('load')] = conv2tag(row[col2num(get_col_ref('load'))].value)
    return thin
        
# Lists 기록함수
def Write_Lists(led_list, inv_list):
    f = open('excel_lists_171108_1.py', 'w')    # 경로설정 / 덮어쓰기
    f.write('led_list = %s\n' % led_list)
    f.write('inv_list = %s\n' % inv_list)
    f.write('\n')
    f.close()

def ts2datetime(ts_str):
    return datetime.datetime.fromtimestamp(int(ts_str)).strftime('%Y/%m/%d-%H:%M:%S')

def datetime2ts(dt_str):
    dt = datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")
    return time.mktime(dt.timetuple())

def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y/%m/%d-%H:%M:%S")

def datetime2str(dt):
    return dt.strftime('%Y/%m/%d-%H:%M:%S')

def add_time(dt_str, days=0, hours=0):
    return datetime2str(str2datetime(dt_str) + datetime.timedelta(days=days, hours=hours))

def is_past(dt_str1, dt_str2):
    return datetime2ts(dt_str1) < datetime2ts(dt_str2)

def is_weekend(dt_str):
    if str2datetime(dt_str).weekday() >= 5: return '1'
    else: return '0'

if __name__ == '__main__':    
    led_list = []
    inv_list = []
    pure_led_list = []
    pure_inv_list = []
    total_num = 1
    name = ''

    wb = load_workbook(filename='ref_list_detail.xlsx', read_only=True)
    ws = wb[u'최종(LED+인버터)']

    #1. extract tags from excel sheet by device type and size
    for row in ws.iter_rows(min_row=4):
        thin = thin_device(row)
        if thin[get_tag_ref('device')] == 'led':
            led_list.append(thin)
        elif thin[get_tag_ref('device')] == 'inverter':
            inv_list.append(thin)    
    print '<총 {0!r}개의 LED와 {1!r}개의 인버터 데이터 추출>\n'.format(len(led_list), len(inv_list))
    
    for led in range(len(led_list)):
        pure_led_list.append(led_list[led]['modem_num'])
        total_num = led+1
    print 'total_led : %s' % total_num
    print 'pure_led_list : %s' % pure_led_list
    #name = 'led'
    #Write_Lists(pure_led_list, name)
    
    for inv in range(len(inv_list)):
        pure_inv_list.append(inv_list[inv]['modem_num'])
        total_num = inv+1
    print 'total_inv : %s' % total_num
    print 'pure_inv_list : %s' % pure_inv_list    
    Write_Lists(pure_led_list, pure_inv_list)

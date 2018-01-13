
# -*- coding: utf-8 -*-
# Author : jeonghoonkang , https://github.com/jeonghoonkang
# Author : jeongmoon417 , https://github.com/jeongmoon417

# 참고 url
# https://code.tutsplus.com/ko/tutorials/how-to-work-with-excel-documents-using-python--cms-25698
# http://egloos.zum.com/mcchae/v/11120944

import datetime
from openpyxl import Workbook
import openpyxl
import time
# import XlsxWriter
import sys
# sys.path.insert(0, '../doc_design')
# (to do) have to find how to add different location directory path and file
# now just using same dir location file

# from openpyxl.workbook import Workbook
# from openpyxl.writer.excel import ExcelWriter
# (error) from openpyxl.cell import get_column_letter
# from openpyxl import load_workbook

global EE_file
global EA_file

class excell_class :
    __ofile = None

    def __init__(self):
        pass

    #@staticmethod
    def open_exc_doc(self,__file):
        # using unicode file name with u syntax
        __ofile = openpyxl.load_workbook(__file)

        #pout = "ix: %s, mdsid: %s, donecheck: %d \r" %(ix, mds_id, donecheck)
        pout = "   ... file opened \n"
        sys.stdout.write(pout)
        sys.stdout.flush()
        return __ofile

    def read_vertical(self, sheet, __start, __end):
        __vertical = []
        #print " ... Please use column[n]:column[m], vertical read "
        cell_of_col = sheet[__start:__end]
        for row in cell_of_col:
            for cell in row:
                v = cell.value
                if v == None:
                    continue # do nothing below code, back to next for loop step
                __vertical.append(v) # 리스트 __vertical에 아이디 추가
        return __vertical #__cnt, __cnt_n # 세로 셀 데이터, 데이터 갯수, None 갯수

'''
    # 입력 리스트를 액셀에 저장
    def save_exc (self, __vdata):
        __t = str(datetime.datetime.now())
        workbook = XlsxWriter.Workbook('takeout_id_result'+__t+'.xlsx')
        worksheet = workbook.add_worksheet()

        row = 0
        col = 0

        for item in (__vdata):
            worksheet.write(row, col, item)
            row += 1
        workbook.close()
'''

def save_vdata(__vdata):
    __t = str(datetime.datetime.now())
    __odata = 'err_id_list='
    print str(__vdata)
    __odata = __odata + str(__vdata)
    print __odata

    filename = '__err_id_input_list_'+__t+'.py'
    __ofile = open(filename,"w")

    __ofile.write(__odata)
    __ofile.close()


# EE에서 수집하는 스마트미터 ID가 '-' 두개 들어가야 하는데,
# 한개만 들어간 경우를 확인하는 함수
# 일반적으로 괜찮은 아이디 00-250060021 (대부분 이런 형식)
# 이상해 보이는 아이디 06-25-0071186 등등
def check_id(__buf):
    __itter = len(__buf) # buf 리스트 길이
    __err_list = []       # 회신할 리스트 버퍼

    for __i in range(__itter) :
        print " check it is OK ? " + __buf[__i]
        try:
            if __buf[__i].index('-',0) != 2 : # 2번 인덱스에 '-' 부재면 get in
                # 이상한 ID 출력, 리스트에 추가
                print __buf[__i]
                __err_list.append(__buf[__i])
            # (2번 인덱스에 '-' 존재하고, 이전 if 에서 결정)
            # 그리고, 5번 인덱스에 '-' 존재이면 get in
            elif __buf[__i].index('-',5) == 5 :
                # 이상한 ID 출력, 리스트에 추가
                print __buf[__i]
                __err_list.append(__buf[__i])
        except:
            #괜찮은 아이디
            print" Proper ID format"# __buf[__i]
            #__err_list.append(__buf[__i])

    return __err_list


def same_site_number(__class, target, inv_led):
    inv_led_sheets = inv_led.get_sheet_names()

    # 인버터 사업장 번호 추출
    inv_sheet = inv_led.get_sheet_by_name(inv_led_sheets[0])
    #print " 확인, INV 여부 ", inv_sheet
    inv_company_nums = __class.read_vertical(inv_sheet, 'b4', 'b101')
    #print inv_branch_nums

    # LED 사업장 번호 추출
    led_sheet = inv_led.get_sheet_by_name(inv_led_sheets[1])
    #print " 확인, LED 여부 ", led_sheet
    led_company_nums = __class.read_vertical(led_sheet, 'b4', 'b416')

    # 통합파일 사업장 번호 추출
    target_sheets = target.get_sheet_names()
    target_sht1 = target.get_sheet_by_name(target_sheets[0])
    target_company_nums = __class.read_vertical(target_sht1, 'a3', 'a561')

    invs = list()
    leds = list()

    # 인버터 사업장 번호 비교
    ibn = len(inv_company_nums)
    tbn = len(target_company_nums)
    #print " ... start loop to  %d times" %(ibn * tbn)

    #same count
    si_cnt = 0

    # 두개 파일에 존재하는 사업장 번호에 대한 미터기 아이디 검색
    # 사업장 번호 같은 경우 (인버터)
    for inv_idx in range(0, ibn):
        for branch_idx in range(0, tbn):
            if str(inv_company_nums[inv_idx]).strip() == str(target_company_nums[branch_idx]).strip():
                si_cnt = si_cnt + 1
                #print " inv same count %d" %(si_cnt)
                same_com_num = target_company_nums[branch_idx]
                # 인버터 ID
                inv_serial = inv_sheet.cell('t'+str(inv_idx+4)).value
                invs.append([same_com_num, u'인버터', inv_serial])

    # 인버터 사업장 번호 비교 결과 저장
    __t = datetime.datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    wb = Workbook()
    ws = wb.active
    file_name = '_del_able_result_same_' + __t + '.xlsx'
    idx = 1

    for item in invs:
        if item[2] is None : continue
        ws["A" + str(idx)] = item[0]
        ws["B" + str(idx)] = item[1]
        ws["C" + str(idx)] = item[2]
        idx += 1

    lbn = len(led_company_nums)

    # 두개 파일에 존재하는 사업장 번호에 대한 미터기 아이디 검색
    # 사업장 번호 같은 경우 (LED)
    for led_idx in range(0, lbn):
        for branch_idx in range(0, tbn):
            # print led[led_idx], branchs[branch_idx]
            if str(led_company_nums[led_idx]).strip() == str(target_company_nums[branch_idx]).strip():
                sl_num = target_company_nums[branch_idx]
                led_serial = led_sheet.cell('s' + str(led_idx+4)).value
                leds.append([ sl_num , u'LED', led_serial])

    # LED 사업장 번호 비교 결과 저장
    # 두개 파일에 동시에 존재하는 경우
    led_wb = Workbook()
    led_ws = wb.active
    for item in leds:
        led_ws["A" + str(idx)] = item[0]
        led_ws["B" + str(idx)] = item[1]
        led_ws["C" + str(idx)] = item[2]
        idx += 1
    wb.save(file_name)
    wb.close()

def diff_site_number(__class, target, inv_led):
    inv_led_sheets = inv_led.get_sheet_names()

    # 인버터 사업장 번호 추출
    inv_sheet = inv_led.get_sheet_by_name(inv_led_sheets[0])
    inv_company_nums = __class.read_vertical(inv_sheet, 'b4', 'b101')

    # LED 사업장 번호 추출
    led_sheet = inv_led.get_sheet_by_name(inv_led_sheets[1])
    led_company_nums = __class.read_vertical(led_sheet, 'b4', 'b416')

    # 통합파일 사업장 번호 추출
    target_sheets = target.get_sheet_names()
    target_sht1 = target.get_sheet_by_name(target_sheets[0])
    target_company_nums = __class.read_vertical(target_sht1, 'a3', 'a561')

    invs = list()
    leds = list()

    # 인버터 사업장 번호 비교
    ibn = len(inv_company_nums)
    tbn = len(target_company_nums)

    #diff count
    diff_cnt = 0

    # 유니코드로 변환
    inv = [unicode(i) for i in inv_company_nums]
    led = [unicode(i) for i in led_company_nums]
    target = [unicode(i) for i in target_company_nums]

    # 두개 파일에 존재하는 사업장 번호 검색
    # KEA 파일에 있는 아이디가 누리텔레콤 파일에 없는 경우 찾음
    # 리스트의 count 함수 이용. 일치하는 멤버없으면 0 리턴, 존재 갯수 리턴
    for idx in range(0, tbn):
        dtect_id = (target[idx])
        ix_inv = inv.count(dtect_id)
        ix_led = led.count(dtect_id)
        # 인버터, LED 리스트에 해당 사업장번호 없음
        if (ix_inv + ix_led) == 0 :
            diff_cnt = diff_cnt + 1
            invs.append([dtect_id, 'is_not_appered_in_NURI_file'])

    # 비교 결과 저장
    __t = datetime.datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    wb = Workbook()
    ws = wb.active
    file_name = '_del_able_result_no_' + __t + '.xlsx'
    idx = 1

    for item in invs:
        #if item[2] is None : continue
        ws["A" + str(idx)] = item[0]
        ws["B" + str(idx)] = item[1]
        idx += 1

    wb.save(file_name)
    wb.close()

if __name__ == "__main__":

    # 입력파일
    EE_file = u"!__161129_16년 EE시장 계량전송장치 입력현황(본사업)_종합본(최종).xlsx"
    EA_file = u"_EE_사업장목록.xlsx"

    cfile = excell_class()
    ee = cfile.open_exc_doc(EE_file)
    ea = cfile.open_exc_doc(EA_file)

    same_site_number(cfile, ea, ee)

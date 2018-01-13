
# -*- coding: utf-8 -*-
# Author : jeonghoonkang , https://github.com/jeonghoonkang
# Author : jeongmoon417 , https://github.com/jeongmoon417

# 참고 url
# https://code.tutsplus.com/ko/tutorials/how-to-work-with-excel-documents-using-python--cms-25698
# http://egloos.zum.com/mcchae/v/11120944

import datetime

import openpyxl
import xlsxwriter
import sys
# sys.path.insert(0, '../doc_design')
# (to do) have to find how to add different location directory path and file
# now just using same dir location file

# from openpyxl.workbook import Workbook
# from openpyxl.writer.excel import ExcelWriter
# (error) from openpyxl.cell import get_column_letter
# from openpyxl import load_workbook

class excell_class :
    __ofile = None

    def __init__(self):
        pass

    #@staticmethod
    def open_exc_doc(self):
        # using unicode file name with u syntax
        __ofile = openpyxl.load_workbook(u"_test__1.xlsx")
        return __ofile

    def read_vertical(self, sheet, __start, __end):
        __vertical = []
        print " ... Please use column[n]:column[m], vertical read "
        cell_of_col = sheet[__start:__end]
        for row in cell_of_col:
            for cell in row:
                v = cell.value
                if v == None:
                    continue # do nothing below code, back to next for loop step
                __vertical.append(v) # 리스트 __vertical에 아이디 추가
        return __vertical #__cnt, __cnt_n # 세로 셀 데이터, 데이터 갯수, None 갯수


    # 입력 리스트를 액셀에 저장
    def save_exc(self, __vdata, __fname):
        __t = str(datetime.datetime.now())
        workbook = xlsxwriter.Workbook(__fname + __t + '.xlsx')
        worksheet = workbook.add_worksheet()

        row = 0
        col = 0

        for item in (__vdata):
            worksheet.write(row, col, item)
            row += 1

        workbook.close()


def save_vdata(__vdata, fname):
    __t = str(datetime.datetime.now())
    __odata = fname + '='

    print str(__vdata)
    __odata = __odata + str(__vdata)

    filename = fname + '.py'
    __ofile = open(filename,"w")
    __ofile.write(__odata)
    __ofile.close()

# EE에서 수집하는 스마트미터 ID가 '-' 두개 들어가야 하는데,
# 한개만 들어간 경우를 확인하는 함수
# 일반적으로 괜찮은 아이디 00-250060021 (대부분 이런 형식)
# 이상해 보이는 아이디 06-25-0071186 등등
def check_id(__buf):
    __itter = len(__buf) # buf 리스트 길이
    __err_list = []       # 회신할 리스트 버퍼

    for __i in range(__itter) :
        print " check it is OK ? " + __buf[__i]
        try:
            if __buf[__i].index('-',0) != 2 : # 2번 인덱스에 '-' 부재면 get in
                # 이상한 ID 출력, 리스트에 추가
                print __buf[__i]
                __err_list.append(__buf[__i])
            # (2번 인덱스에 '-' 존재하고, 이전 if 에서 결정)
            # 그리고, 5번 인덱스에 '-' 존재이면 get in
            elif __buf[__i].index('-',5) == 5 :
                # 이상한 ID 출력, 리스트에 추가
                print __buf[__i]
                __err_list.append(__buf[__i])
        except:
            #괜찮은 아이디
            print" Proper ID format"# __buf[__i]
            #__err_list.append(__buf[__i])

    return __err_list

if __name__ == "__main__":

# open excell file
    eclass = excell_class()
    op = eclass.open_exc_doc()
    sheets = op.get_sheet_names()
    print " sheets = ", sheets

    sh1 = op.get_sheet_by_name(sheets[0])
    print " name =", sh1
    buf = eclass.read_vertical(sh1,'b1','b541')
    save_vdata(buf,"result_ids")
    eclass.save_exc(buf,"result_ids")

    exit (" ...congrats, finish")

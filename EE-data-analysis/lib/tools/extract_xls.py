import sys
import os
import getopt
import logging
import copy
import pickle
import json

import pip
installed_packages = [package.project_name for package in pip.get_installed_distributions()]
requested_packages = ['openpyxl', 'numpy']
for package in requested_packages:
    if package in installed_packages:
        pass
    else:
        print(package + ' must be already installed to use extract_xls.py')
        sys.exit(1)
from openpyxl import load_workbook
import numpy as np


class xls2list:
    def __init__(self, ):
        self.column_ref = None
        self.tag_ref = None
        self.wb = None
        self.ws = None
        self.column_name = { # Refer to column names in tsdb_tool.py
            'device_serial':    '_mds_id',
            'modem_serial':     'modem_num',
            'holiday':          'holiday',
            'company':          'company',
            'device':           'device_type',
            'building':         'building',
            'load':             'load',
            'size':             'size',
            'business':         'business',
            'device_type':      'device_type',
            'serial':           '_mds_id',
            'modem_num':        'modem_num',
            'led_serial':       '_mds_id',
            'led_modem_serial': 'modem_num',
            'led_size':         'size',
            'inv_serial':       '_mds_id',
            'inv_modem_serial': 'modem_num',
            'inv_size':         'size',
            }

    def column(self, str):
        if str in self.column_name:
            return self.column_name[str]
        else:
            return ''

    def col2num(self, col_str):
        expn = 0
        col_num = 0
        for char in reversed(col_str):
            col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
            expn += 1
        return col_num - 1

    def conv2tag(self, val_str):
        if val_str in self.tag_ref:
            return self.tag_ref[val_str]
        else:
            return 'none'

    def get_col_ref(self, key):
        if key in self.column_ref:
            return self.column_ref[key][0]
        else:
            return 'ZZZ'

    def get_tag_ref(self, key):
        if key in self.column_ref:
            return self.column_ref[key][1]
        else:
            return 'unknown'

    def thin_device(self, row):
        thin = {self.get_tag_ref('device'): 'none'}
        if row[self.col2num(self.get_col_ref('led_modem_serial'))].value != None:
            if row[self.col2num(self.get_col_ref('led_size'))].value != None:
                thin[self.get_tag_ref('device')] = self.conv2tag(row[self.col2num(self.get_col_ref('device'))].value)
                thin[self.get_tag_ref('led_serial')] = row[self.col2num(self.get_col_ref('led_serial'))].value.encode('utf8')
                thin[self.get_tag_ref('led_modem_serial')] = row[self.col2num(self.get_col_ref('led_modem_serial'))].value.encode('utf8')
                thin[self.get_tag_ref('company')] = self.conv2tag(row[self.col2num(self.get_col_ref('company'))].value)
                thin[self.get_tag_ref('building')] = self.conv2tag(row[self.col2num(self.get_col_ref('building'))].value)
                thin[self.get_tag_ref('business')] = self.conv2tag(row[self.col2num(self.get_col_ref('business'))].value)
                #thin[self.get_tag_ref('business')] = 'none'
                thin[self.get_tag_ref('load')] = self.conv2tag(row[self.col2num(self.get_col_ref('load'))].value)
        elif row[self.col2num(self.get_col_ref('inv_modem_serial'))].value != None:
            if row[self.col2num(self.get_col_ref('inv_size'))].value != None:
                thin[self.get_tag_ref('device')] = self.conv2tag(row[self.col2num(self.get_col_ref('device'))].value)
                thin[self.get_tag_ref('inv_serial')] = row[self.col2num(self.get_col_ref('inv_serial'))].value.encode('utf8')
                thin[self.get_tag_ref('inv_modem_serial')] = row[self.col2num(self.get_col_ref('inv_modem_serial'))].value.encode('utf8')
                thin[self.get_tag_ref('company')] = self.conv2tag(row[self.col2num(self.get_col_ref('company'))].value)
                thin[self.get_tag_ref('building')] = self.conv2tag(row[self.col2num(self.get_col_ref('building'))].value)
                thin[self.get_tag_ref('business')] = self.conv2tag(row[self.col2num(self.get_col_ref('business'))].value)
                #thin[self.get_tag_ref('business')] = 'none'
                thin[self.get_tag_ref('load')] = self.conv2tag(row[self.col2num(self.get_col_ref('load'))].value)
        return thin

    def load(self, input_file, sheet):
        try:
            self.wb = load_workbook(filename=input_file, read_only=True)
        except:
            return None

        self.ws = self.wb.worksheets[sheet]

        return self.ws

    def convert(self, column_ref, tag_ref):
        self.column_ref = column_ref
        self.tag_ref = tag_ref
        led_list = []
        inv_list = []

        if not self.ws:
            return {}

        #Extract tags from excel sheet by device type and size
        for row in self.ws.iter_rows(min_row=4):
            thin = self.thin_device(row)
            if thin[self.get_tag_ref('device')] == 'led':
                led_list.append(thin)
            elif thin[self.get_tag_ref('device')] == 'inverter':
                inv_list.append(thin)

        return {'leds': led_list, 'inverters': inv_list}

    def decode(self, device):
        tmp = copy.deepcopy(device)
        if tmp['device_type'] == 'led':
            tmp[self.get_tag_ref('led_serial')] = tmp[self.get_tag_ref('led_serial')].decode('utf-8')
            tmp[self.get_tag_ref('led_modem_serial')] = tmp[self.get_tag_ref('led_modem_serial')].decode('utf-8')
        elif tmp['device_type'] == 'inverter':
            tmp[self.get_tag_ref('inv_serial')] = tmp[self.get_tag_ref('inv_serial')].decode('utf-8')
            tmp[self.get_tag_ref('inv_modem_serial')] = tmp[self.get_tag_ref('inv_modem_serial')].decode('utf-8')
        else:
            return {}
        return tmp

def help(err):
    print('usage:')
    print('\t-i iii : specify the input xls file')
    print('\t-s sss : specify the position of the sheet')
    print('\t-o ooo : specify the output dump file')
    print('\t-d     : turn on debug logging')
    print('\tex) python extract_xls.py -i ref_list.xlsx -s 2 -o list.json -d')
    sys.exit(err)

if __name__ == '__main__':
    if sys.version_info < (3,6,0):
        sys.stderr.write('You need python 3.6 or later to run this script.\n')
        sys.exit(1)

    input_file = ''
    output_file = ''
    sheet = 0
    log_level = logging.INFO
    json_output = True

    try:
        opts, args = getopt.getopt(sys.argv[1:], 'i:s:o:d')
    except getopt.GetoptError as err:
        print(err)
        help(2)

    if opts:
        for o, a in opts:
            if o == '-i':
                input_file = a
            elif o == '-s':
                sheet = a
            elif o == '-o':
                output_file = a
            elif o == '-d':
                log_level = logging.DEBUG
            else:
                help(1)

    if not input_file:
        help(1)

    logger = logging.getLogger(__name__)
    logger.setLevel(log_level)
    ch = logging.StreamHandler()
    ch.setLevel(log_level)
    ch.setFormatter(logging.Formatter('%(asctime)s %(levelname)s:%(message)s'))
    logger.addHandler(ch)

    # Rearrange columns to match your sheet
    column_ref = {
        'company': ['E', 'company'],
        'device': ['F', 'device_type'],
        'building': ['M', 'building'],
        'load': ['W', 'load'],
        'led_serial': ['S', '_mds_id'],
        'led_modem_serial': ['T', 'modem_num'],
        'led_size': ['R', 'size'],
        'inv_serial': ['AB', '_mds_id'],
        'inv_modem_serial': ['AC', 'modem_num'],
        'inv_size': ['Z', 'size'],
        'business': ['AL', 'business'],
        }

    tag_ref = {
        u'공장': 'factory', u'일반건물': 'office', u'마트': 'mart', u'공동주택': 'apt', u'대형마트': 'mart', u'기타': 'etc',  # 건물용도
        u'대기업': 'bigcompany', u'중견': 'middlecompany', u'중소': 'smallcompany', u'비영리법인': 'etc', u'해당없음': 'etc',
        u'병원': 'hospital',  # 사업장구분
        u'금속': 'metal', u'산업기타': 'industryEtc', u'건물기타': 'building_etc', u'제지목재': 'wood', u'요업': 'ceramic',
        u'백화점': 'departmentStore', u'섬유': 'fiber', u'학교': 'school', u'식품': 'food', u'화공': 'ceramic',
        u'상용': 'sangyong',  # 업종별-세부용도
        u'LED': 'led', u'인버터': 'inverter', u'형광등': 'heungwanglamp', u'메탈할라이드': 'metalhalide',  # 품목
        np.NaN: 'none', u'블로워': 'blower', u'컴프래서': 'compressor', u'펌프': 'pump', u'팬': 'fan',
        }

    xls = xls2list()
    ws = xls.load(input_file, int(sheet))
    if not ws:
        logger.error('Input file is invalid.')
        sys.exit(1)

    device_list = xls.convert(column_ref, tag_ref)

    logger.info('Input file: %s, Selected sheet: %s, Output file: %s', input_file, ws.title, output_file)
    logger.info('Total {0!r} of LEDs and {1!r} of Inverters extracted.'.format(len(device_list['leds']), len(device_list['inverters'])))

    if log_level == logging.DEBUG:
        logger.debug('led list:')
        for led in device_list['leds']:
            logger.debug(json.dumps(xls.decode(led)))
        logger.debug('inverter list:')
        for inv in device_list['inverters']:
            logger.debug(json.dumps(xls.decode(inv)))

    if output_file:
        if json_output:
            json_list = { 'leds': [], 'inverters': []}
            for led in device_list['leds']:
                json_list['leds'].append(xls.decode(led))
            for inv in device_list['inverters']:
                json_list['inverters'].append(xls.decode(inv))
            with open(output_file, 'w') as f:
                json.dump(json_list, f, ensure_ascii=False)
        else:
            pickle.dump(device_list, open(output_file, 'wb'))
            #tmp = pickle.load(open(output_file, 'rb'))

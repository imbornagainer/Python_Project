# -*- coding: utf-8 -*-

import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import json
import socket
import copy
from openpyxl import load_workbook
import numpy as np
import datetime
import time
import gzip
import shutil
import openpyxl

# 해당 문서의 sheet 탐색 - get_sheet_names()
# 해당 문서의 다수 sheet 접근 - get_sheet_by_name()
if __name__ == '__main__':

    l_w = load_workbook('ref_list_kjh.xlsx')
    l_ws = l_w[u'최종(LED+인버터)']
    
    for row in l_ws.iter_rows(min_row=4, max_row=539, max_col=5):   # 행을 기준으로 iter(반복문)
        for cell in row:
            print "cell : %s" % (cell)
            print "cell.value : %s" % (cell.value)
    
    for col in l_ws.iter_cols(min_row=4, max_row=539, max_col=5): # 열을 기준으로 iter(반복문) - min_row - 최소 시작할(행), max_row - 마지막(행), max_col - 마지막(열) 
        for cell in col:
            print "cell : %s" % (cell)
            print "cell.value : %s" % (cell.value)    